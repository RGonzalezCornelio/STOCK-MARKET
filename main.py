#MAIN

import openpyxl
import requests
from bs4 import BeautifulSoup
import re
import json
import signal
from colorama import init, Fore


#Le meteremos un handler seguramente porque son mas de 5000 lineas
def handler(signum, frame):
    with open("./counter.txt", 'w') as file:
        file.write(str(x))
    work.save("./STOCK MARKET NAMES.xlsx")
    print(Fore.RED + "❌")
    exit(1)

init()
signal.signal(signal.SIGINT, handler)

work = openpyxl.load_workbook("./STOCK MARKET NAMES.xlsx")
companies = work['NAMES']
column = 1 #Columns with the company names
y = int(open('./counter.txt').readline())
linkstarter = "https://finance.yahoo.com/quote/"
headers = {"User-Agent":"Mozilla/5.0"}

for x in range(y, 5582):

    print(Fore.RESET + "processing "+ str(x) + "... ", end= '')
    jsonf = {} #Guardaremos la informacion en un json
    cell = companies.cell(row=x, column=column)
    stock_name = str(companies.cell(row=x, column=1).value)

    #Hay stock names que tienen '^' en el nombre, vamos a saltarlos porque son depositos y no interesan
    if '^' not in stock_name:

        jsonf["Stock Name"] = stock_name

        link = linkstarter + stock_name + "/key-statistics/"
        jsonf["link"] = link
        #print(company_name)
        print(stock_name + ", " + link)

        try:
            respuesta = requests.get(link, headers=headers)
            if str(respuesta) != "<Response [200]>":
                print("Error request")
                print(Fore.RED + "❌")
                continue
        except Exception:
            handler("", "")

        
        #Empezamos a coger datos de la pagina web
        data = BeautifulSoup(respuesta.text, 'html.parser')

        company = str(data.select('title')[0]).split(">")[1].split('(')[0]
        #print(company)
        #Esto se trata de un error en la web, supongo que sera por el gran numero de peticiones que se hacen a la pagina web. En todo caso haremos que el programa pare.
        if company == "Symbol Lookup from Yahoo Finance</title":
            
            print("Error company name")
            handler("", "")
        
        jsonf["Company Name"] = company
        

        try:
            stock_value = str(data.find('div', class_="container yf-1tejb6")).split('>')[3].split('<')[0]
        except Exception:
            stock_value = 0

        jsonf["Stock Value"] = stock_value
        print("Stock Value: " + str(stock_value))

        #statictics_table = str(data.find('table', class_="table yf-kbx2lo"))
        #statictics_table_values = str(data.find_all('td', class_="yf-kbx2lo"))
        #statictics_table_values = statictics_table.find('table', class_="table yf-kbx2lo")
        #print(statictics_table_values)

        try:
            market_cap = str(data.find_all('td', class_="yf-kbx2lo")[1]).split('>')[1].split('<')[0]
        except Exception:
            market_cap = 0

        try:
            enterprise_value = str(data.find_all('td', class_="yf-kbx2lo")[8]).split('>')[1].split('<')[0]
        except Exception:
            enterprise_value = 0

        try:
            trailing_PE = str(data.find_all('td', class_="yf-kbx2lo")[15]).split('>')[1].split('<')[0]
        except Exception:
            trailing_PE = 0
        try:
            forward_PE = str(data.find_all('td', class_="yf-kbx2lo")[22]).split('>')[1].split('<')[0]
        except Exception:
            forward_PE = 0

        try:
            enterprise_value_EBITDA = str(data.find_all('td', class_="yf-kbx2lo")[57]).split('>')[1].split('<')[0]
        except Exception:
            enterprise_value_EBITDA = 0
        #print("market_cap: " + market_cap + "Enterprise value: " + enterprise_value + "Trailing P/E: " + trailing_PE + "forward_PE: " + forward_PE + "enterprise_value_EBITDA: " + enterprise_value_EBITDA)

        jsonf["Market Cap"] = market_cap
        jsonf["Enterprise Value"] = enterprise_value
        jsonf["Trailing P/E"] = trailing_PE
        jsonf["Forward P/E"] = forward_PE
        jsonf["Enterprise Value/EBITDA"] = enterprise_value_EBITDA


        try:
            profit_margin = str(data.find_all('td', class_="value yf-vaowmx")[2]).split('>')[1].split('<')[0]
        except Exception: 
            profit_margin = 0

        try:
            operating_margin = str(data.find_all('td', class_="value yf-vaowmx")[3]).split('>')[1].split('<')[0]
        except Exception: 
            operating_margin = 0

        try:
            quarterly_revenue_growth = str(data.find_all('td', class_="value yf-vaowmx")[8]).split('>')[1].split('<')[0]
        except Exception: 
            quarterly_revenue_growth = 0
        
        try:
            quarterly_earnings_growth = str(data.find_all('td', class_="value yf-vaowmx")[13]).split('>')[1].split('<')[0]
        except Exception: 
            quarterly_earnings_growth = 0

        try:
            beta = str(data.find_all('td', class_="value yf-vaowmx")[22]).split('>')[1].split('<')[0]
        except Exception: 
            beta = 0

        try:
            percentage_by_insiders = str(data.find_all('td', class_="value yf-vaowmx")[34]).split('>')[1].split('<')[0]
        except Exception: 
            percentage_by_insiders = 0

        try:
            forward_annual_divident_rate = str(data.find_all('td', class_="value yf-vaowmx")[41]).split('>')[1].split('<')[0]
        except Exception: 
            forward_annual_divident_rate = 0

        try:
            five_year_divident_rate = str(data.find_all('td', class_="value yf-vaowmx")[45]).split('>')[1].split('<')[0]
        except Exception: 
            five_year_divident_rate = 0
        #print("profit_margin: " + profit_margin + " operating_margin: " + operating_margin + " quarterly_revenue_growth: " + quarterly_revenue_growth + " quarterly_earnings_growth: " + quarterly_earnings_growth\
        #       + " beta: " + beta + " percentage_by_insiders: " + percentage_by_insiders + " forward_annual_divident_rate" + forward_annual_divident_rate + " five_year_divident_rate: " + five_year_divident_rate)

        jsonf["Profit Margin"] = profit_margin
        jsonf["Operating Margin"] = operating_margin
        jsonf["Quarterly Revenue Growth"] = quarterly_revenue_growth
        jsonf["Quarterly Earnings Growth"] = quarterly_earnings_growth
        jsonf["Beta (5Y Monthly)"] = beta
        jsonf["Percentage by insiders"] = percentage_by_insiders
        jsonf["Forward Annual Dividend Rate"] = forward_annual_divident_rate
        jsonf["5 Year Average Dividend Yield"] = five_year_divident_rate
        

        jsonfile = open('./DataJSON/' + str(x) + "_" + stock_name + '.json', 'w')
        json.dump(jsonf, jsonfile)
        print(Fore.GREEN + "✓")

work.save("./STOCK MARKET NAMES.xlsx")